# **************** PARTE 1: EN ESTA PARTE SE CONSTRUYE LA RED NEURONAL CONVOLUCIONAL *****************

# Se importan las librerias a utilizar
from keras.models import Sequential
from keras.layers import Flatten, Dense, Dropout
from keras.layers.convolutional import Conv2D, MaxPooling2D
from keras.preprocessing.image import ImageDataGenerator
from keras import backend as K
import os
import numpy as np
import matplotlib.pyplot as plt
import cv2 
from openpyxl import Workbook

K.clear_session() #limpia el backend de keras para que no haya informacion basura 

# Se definen las rutas de los conjuntos de entrenamiento y el de validacion
train_path = 'C:/Users/saulm/Documents/python/deep_learning/cnn/Coral_Reef_Disease/DATASET/TRAINING_SET'
test_path = 'C:/Users/saulm/Documents/python/deep_learning/cnn/Coral_Reef_Disease/DATASET/TEST_SET'


#Leyendo imagenes para comprobar que si se puede acceder a la carpeta de entrenamiento
'''enf_path = 'C:/Users/saulm/Documents/python/deep_learning/cnn/Coral_Reef_Disease/DATASET/TRAINING_SET/ENFERMOS/'
enf_training = []
cont = 0

for imgn in os.listdir(enf_path):
  cont+=1
  imagen = cv2.imread(os.path.join(enf_path, imgn))
  enf_training.append(imagen)

enf_training = np.array(enf_training, dtype=object)
print(enf_training.shape)
print(f'El total de imagenes es: {cont}')
print(np.array(enf_training[972]).shape)
plt.figure()
plt.imshow(np.squeeze(enf_training[972]))
plt.colorbar()
plt.grid(False)
plt.show()'''


# Debido a la poca cantidad de imagenes de muestra, se utilzan GENERADORES los cuales tienen la funcion de 
# a partir de una imagen usar transformaciones geometricas para aumentar el conjunto de entrenamiento
# las transformaciones solo aplican para el conjunto de entrenamiento, en el de validacion solo se convierte a valor flotante
train_datagen = ImageDataGenerator(
        rescale=1./255,
        shear_range=0.3,
        zoom_range=0.3,
        horizontal_flip=True)

test_datagen = ImageDataGenerator(rescale=1./255)

training_set = train_datagen.flow_from_directory(
        train_path,
        target_size=(200, 200),
        batch_size=32,
        class_mode='binary')


test_set = test_datagen.flow_from_directory(
        test_path,
        target_size=(200, 200),
        batch_size=32,
        class_mode='binary')

print(test_set.class_indices) # Para saber que etiquetas tienen las clases [0-Enfermo] o [1-Sano] 
print(training_set.class_indices)

# ==== Calculando los pasos de Entrenamiento y validacion ====
# los pasos se calcula diviendo el tamaño de cada conjunto entre el tamaño de lote
pasos_train = training_set.n//training_set.batch_size
pasos_val = test_set.n//test_set.batch_size


# Creando la Red Neuronal Convolucional
model = Sequential()

# Step - Convolution
# Si la imagen es mayor a 128px usar un kernel size mayor a 3x3
model.add(Conv2D(filters=32, kernel_size=5, strides=(1,1), padding='same', activation='relu', input_shape=(200,200,3)))
model.add(MaxPooling2D(pool_size = (2, 2)))

# Adding a second convolutional layer
model.add(Conv2D(filters=64, kernel_size=3, strides=(1,1), padding='same', activation='relu'))
model.add(MaxPooling2D(pool_size = (2, 2)))

# Step - Flattening
model.add(Flatten()) # Una vez teniendo nuestros mapas de de caracteristicas, los convertimos a un vector lineal, estos datos seran las entradas a la red perceptron

# Step - Full connection, en este paso se crea la red tipo perceptron con 128 neuronas
model.add(Dense(units = 128, activation = 'relu'))
""" EL Overfitting es cuando el modelo solo toma como validos los datos con los cuales se entreno
    y no logra generalizar con datos que sean diferentes a la base de datos inicial"""
model.add(Dropout(0.5))   # El dropout establece aleatoriamente entradas en 0 en cada paso de entrenamiento lo que ayuda a evitar el sobreajuste

#Se establece una sola neurona para la capa de salida en vez de dos, ya que se necesitan menos parametros y calculos
#La salida 0 (<0.5) se considera clase ENFERMO y 1 (>=0.5) se considera clase SANO
#Se utiliza la funcion de activacion sigmoidal o logistica, ya que esta funcion existe entre 0 y 1
#en donde cruzan asintotas horizontales que caracterizan a esta funcion, ademas de que tiene una derivada sencilla
#por lo tanto es usada en modelos donde tengamos que predecir una probabilidad entre 0 y 1
model.add(Dense(units = 1, activation = 'sigmoid'))

# Compiling the CNN
# La funcion de perdida nos informa que tan precisa es nuestra red
# La pérdida de entropía cruzada se utiliza al ajustar los pesos del modelo durante el entrenamiento. 
# El objetivo es minimizar la pérdida, es decir, cuanto menor sea la pérdida, mejor será el modelo.
# Calcula la pérdida de entropía cruzada entre etiquetas verdaderas y etiquetas predichas.
model.compile(optimizer = 'adam', loss = 'binary_crossentropy', metrics = ['accuracy'])

# Part 2 - Fitting the CNN to the images
# se le asignan los parametros al modelo para que pueda entrenarse y se almacenan en la variable history
history = model.fit(training_set, steps_per_epoch=pasos_train, epochs=15, validation_data=test_set, validation_steps=pasos_val)


# --- Creacion del directorio donde se guardara el modelo ---
directorio ='./MODELO/'
# si el directorio no existe se crea
if not os.path.exists(directorio):
    os.mkdir(directorio)  # se crea la carpeta en la ruta actual del proyecto
#el modelo y los pesos son guardados respectivamente
model.save('./MODELO/MODELO_V1.h5')
model.save_weights('./MODELO/PESOS_V1.h5')

# Con el objeto creado podemos acceder al diccionario donde se almacenan los parametros que arroja el modelo
history_dict = history.history
print(history_dict.keys())

# Estos son las claves del diccionario obtenido de history_dict.keys(), con ellos obtenemos los valores de cada clave que se generaron por cada epoca de entrenamiento
acc      = history.history[     'accuracy' ]
val_acc  = history.history[ 'val_accuracy' ] # esta es la presicion FINAL!
loss     = history.history[    'loss' ]
val_loss = history.history['val_loss' ]  

# Procedemos a mostrar los datos obtenidos durante el entrenamiento
print(acc)  
print(val_acc) 
print(loss)
print(val_loss) 

print("----------------- RESULTADO DEL MODELO -----------------")
print()
print()
print()
print()
print("La Presición FINAL del modelo es: ", history_dict['val_accuracy'][-1]) # mostramos el valor de la presicion almacenado en la ultima posicion de la lista del valor en la key val_accuracy

print("----------------- GRÁFICAS -----------------")
print()
print()
print()
print()

epochs = range(1, len(loss) + 1)
plt.plot(epochs, loss, 'y', label='Training loss')
plt.plot(epochs, val_loss, 'r', label='Validation loss')
plt.title('Training and validation loss')
plt.xlabel('Epochs')
plt.ylabel('Loss')
plt.legend()
plt.show()

plt.plot(epochs, acc, 'y', label='Training acc')
plt.plot(epochs, val_acc, 'g', label='Validation acc')
plt.title('Training and validation accuracy')
plt.xlabel('Epochs')
plt.ylabel('Accuracy')
plt.legend()
plt.show()

# Aqui se crea el archivo excel donde se almacena los datos del entrenamiento
wb = Workbook()

hoja = wb.active
hoja.title = "registro del Entrenamiento"

num=0
fila=2
accuary=2
val_ac=3
val_l=4
lss = 5

for accu, vala, vals, loss in zip(acc, val_acc, val_loss, loss):
    hoja.cell(row=1, column=2, value='Accuracy')
    hoja.cell(row=1, column=3, value='Val_accuracy')
    hoja.cell(row=1, column=4, value='Val_loss')
    hoja.cell(row=1, column=5, value='Loss')
    
    hoja.cell(column=1, row=fila, value=num)
    hoja.cell(column=accuary, row=fila, value= accu)
    hoja.cell(column=val_ac, row=fila, value= vala)
    hoja.cell(column=val_l, row=fila, value= vals)
    hoja.cell(column=lss, row=fila, value= loss)
    fila+=1
    num+=1
wb.save('./MODELO/RESULTADOS.xlsx') # los datos son guardados en la ruta especificada